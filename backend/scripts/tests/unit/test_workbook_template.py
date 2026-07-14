"""The generated template must be importable by the parser that consumes it.

These tests are the guard against the template and the parser drifting apart —
a renamed header or sheet in either one breaks the round trip here rather than
silently dropping a column on a real upload.
"""

import io

import openpyxl
import pytest

from app.features.imports.parsers.workbook_parser import (
    HEADER_ALIASES,
    SHEET_CONFIG,
    _clean_header,
    parse_workbook,
)
from app.features.imports.template import (
    RESUMES_LINK_ROW,
    TEMPLATE_SHEETS,
    build_workbook_template,
)


@pytest.fixture(scope="module")
def template_bytes() -> bytes:
    return build_workbook_template()


def test_template_parses_without_errors(template_bytes: bytes):
    parsed = parse_workbook(template_bytes)

    errors = [i for i in parsed.issues if i.severity == "error"]
    assert errors == [], f"Template produced parse errors: {errors}"

    # One example row per sheet.
    assert len(parsed.students) == 1
    assert len(parsed.mentors) == 1
    assert len(parsed.mentor_projects) == 1
    assert len(parsed.probable_projects) == 1


def test_template_example_row_populates_every_student_column(template_bytes: bytes):
    parsed = parse_workbook(template_bytes)
    _, student, _ = parsed.students[0]

    # Every column the template advertises must survive the round trip. The
    # developer-profile columns in particular are only read if they are listed
    # in SHEET_CONFIG["Students Info"]["optional"].
    assert student.name == "Jane Doe"
    assert student.registration_number == "MDS2025001"
    assert student.email == "jane.doe@example.com"
    assert student.phone == "9876543210"
    assert student.skills == "Python, PyTorch, SQL"
    assert student.github_username == "janedoe"
    assert student.leetcode_username == "janedoe"
    assert student.codeforces_username == "janedoe"
    assert student.kaggle_username == "janedoe"
    assert student.scholar_id == "AbCdEfGhIjK"
    assert student.live_project_links == "https://jane-doe.example.com"


def test_template_headers_match_sheet_config(template_bytes: bytes):
    """Every column the parser expects has a header in the template, and the
    sheet names are ones the parser recognises."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))

    for sheet_name, headers in TEMPLATE_SHEETS.items():
        config = SHEET_CONFIG[sheet_name]
        cleaned = [_clean_header(h) for h in headers]

        for col in config["required"] + config["optional"]:
            aliases = set(HEADER_ALIASES.get(col, [col]))
            assert any(
                h in aliases for h in cleaned
            ), f"'{sheet_name}' template is missing a column for '{col}'"

        assert _clean_header(sheet_name) in config["aliases"]
        assert sheet_name in wb.sheetnames


def test_template_leaves_the_resume_link_cell_empty(template_bytes: bytes):
    """A placeholder Drive URL here would be picked up as the real folder and
    fail the import, so the cell must ship blank."""
    parsed = parse_workbook(template_bytes)
    assert parsed.resumes_url is None

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    students = wb["Students Info"]
    assert students.cell(row=RESUMES_LINK_ROW, column=1).value == "All Resumes"
    assert students.cell(row=RESUMES_LINK_ROW, column=2).value is None
