"""Generates the blank workbook that users fill in and upload.

The headers here must stay parseable by ``workbook_parser``: every header is
matched against ``HEADER_ALIASES`` after ``_clean_header`` lowercases it, and
every sheet name against ``SHEET_CONFIG`` aliases. ``test_workbook_template``
asserts a round-trip so the two cannot drift apart.
"""

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# The row the parser reads the shared Drive folder hyperlink from.
RESUMES_LINK_ROW = 25

TEMPLATE_SHEETS: dict[str, list[str]] = {
    "Students Info": [
        "Name",
        "Registration Number",
        "Email",
        "Phone",
        "File",
        "Skills",
        "GitHub Username",
        "LeetCode Username",
        "Codeforces Username",
        "Kaggle Username",
        "Scholar ID",
        "Live Project Links",
    ],
    "Mentors info": [
        "Mentors",
        "Email ID",
    ],
    "Mentors-projects": [
        "Mentors",
        "Short Profile of the Mentor",
        "Project Title",
        "Project Abstract",
        "Pre-requisites",
        "Student's Preference - 1",
        "Student's Preference - 2",
        "Student's Preference - 3",
        "Selected Students (to be filled by mentors)",
    ],
    "Probable projects": [
        "Project Idea",
        "Author",
        "Topic",
    ],
}

EXAMPLE_ROWS: dict[str, list[str]] = {
    "Students Info": [
        "Jane Doe",
        "MDS2025001",
        "jane.doe@example.com",
        "9876543210",
        "jane-doe-resume.pdf",
        "Python, PyTorch, SQL",
        "janedoe",
        "janedoe",
        "janedoe",
        "janedoe",
        "AbCdEfGhIjK",
        "https://jane-doe.example.com",
    ],
    "Mentors info": [
        "Dr. John Smith",
        "john.smith@example.com",
    ],
    "Mentors-projects": [
        "Dr. John Smith",
        "Associate Professor, works on graph neural networks.",
        "Fraud detection with graph neural networks",
        "Build a GNN that flags fraudulent transactions in a payments graph.",
        "Python, PyTorch, Graph Theory",
        "MDS2025001",
        "MDS2025002",
        "",
        "",
    ],
    "Probable projects": [
        "On-device speech recognition for low-resource languages",
        "Jane Doe",
        "Speech / NLP",
    ],
}

# Column widths, keyed by header, so the generated file is readable as-is.
_WIDE_COLUMNS = {
    "Short Profile of the Mentor": 46,
    "Project Abstract": 60,
    "Project Idea": 60,
    "Pre-requisites": 34,
    "Skills": 34,
    "Live Project Links": 34,
    "Selected Students (to be filled by mentors)": 34,
    "Email": 28,
    "Email ID": 28,
    "Project Title": 40,
}

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HINT_FONT = Font(italic=True, color="6B7280")

_INSTRUCTIONS = [
    "How to fill in this workbook",
    "",
    "1. Keep the four sheet names and the header row exactly as they are — the "
    "importer matches on them.",
    "2. Row 2 of each sheet is an example. Replace it with your own data or "
    "delete it before uploading.",
    "3. Add as many rows as you need. Blank rows are ignored.",
    "4. Leave a cell blank when you have no value for it — do not write 'N/A'.",
    "",
    "Students Info",
    "  · Name and Registration Number are required; everything else is optional.",
    "  · Registration Number must be unique within the workbook.",
    "  · Skills: comma-separated (e.g. Python, PyTorch, SQL).",
    "  · The username columns take the username only, not the full profile URL.",
    f"  · Cell B{RESUMES_LINK_ROW} (next to the 'All Resumes' label) is where you "
    "paste the shared folder link holding every student's resume PDF. Name each "
    "PDF after the student so resumes can be matched to rows.",
    "",
    "Mentors info",
    "  · Mentors (the mentor's name) is required.",
    "",
    "Mentors-projects",
    "  · Mentors and Project Title are required.",
    "  · The mentor name must match a row in 'Mentors info' exactly.",
    "  · Preference and Selected Students columns take registration numbers.",
    "",
    "Probable projects",
    "  · Optional sheet. Project Idea is required if you use it.",
]


def _write_sheet(ws: Worksheet, headers: list[str], example: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = _WIDE_COLUMNS.get(
            header, 22
        )

    for idx, value in enumerate(example, start=1):
        ws.cell(row=2, column=idx, value=value)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def build_workbook_template() -> bytes:
    """Build the blank import workbook and return it as .xlsx bytes."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    for sheet_name, headers in TEMPLATE_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        _write_sheet(ws, headers, EXAMPLE_ROWS[sheet_name])

    # The parser skips the row labelled "All Resumes" when reading students and
    # reads the shared resume-folder hyperlink from column B of it. Leave the
    # link cell blank: any placeholder URL here would be picked up as the real
    # folder and fail the import.
    students = wb["Students Info"]
    students.cell(row=RESUMES_LINK_ROW, column=1, value="All Resumes")
    hint = students.cell(
        row=RESUMES_LINK_ROW,
        column=3,
        value="← paste the shared resume-folder link in column B",
    )
    hint.font = _HINT_FONT

    instructions = wb.create_sheet("Instructions", 0)
    instructions.column_dimensions["A"].width = 100
    for idx, line in enumerate(_INSTRUCTIONS, start=1):
        cell = instructions.cell(row=idx, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    instructions["A1"].font = Font(bold=True, size=14)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
