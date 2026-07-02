import io
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.features.imports.schemas import (
    MentorProjectRow,
    MentorRow,
    ParsedWorkbook,
    ProbableProjectRow,
    StudentRow,
    ValidationIssueOut,
)


def _clean_header(header: str) -> str:
    """Normalize headers by removing whitespace and newlines, and lowercasing."""
    if not isinstance(header, str):
        return ""
    return header.strip().replace("\n", "").replace("\r", "").lower()


# Map canonical header names to their expected variations in the workbook.
HEADER_ALIASES = {
    # Students Info
    "name": ["name", "student name"],
    "registration number": [
        "registration number",
        "registration_number",
        "registration no",
        "registration no.",
        "reg no",
        "roll number",
    ],
    "email": ["email", "email id", "mail id"],
    "phone": ["phone", "phone number", "contact", "contact number"],
    "file": ["file", "resume", "resume file", "cv"],
    # Mentors info
    "mentors": ["mentors", "mentor name", "mentor", "faculty name"],
    "email id": ["email", "email id", "mail id"],
    # Mentors-projects
    "short profile of the mentor": [
        "short profile of the mentor",
        "mentor profile",
        "short profile",
    ],
    "project title": ["project title", "title"],
    "project abstract": ["project abstract", "abstract", "description"],
    "pre-requisites": ["pre-requisites", "prerequisites", "pre requisites"],
    "student's preference - 1": ["student's preference - 1", "student preference 1"],
    "student's preference - 2": ["student's preference - 2", "student preference 2"],
    "student's preference - 3": ["student's preference - 3", "student preference 3"],
    "selected students": [
        "selected students (to be filled by mentors)",
        "selected students",
    ],
    "github username": ["github username", "github", "github profile"],
    "leetcode username": ["leetcode username", "leetcode", "leetcode profile"],
    "codeforces username": [
        "codeforces username",
        "codeforces",
        "codeforces profile",
    ],
    "kaggle username": ["kaggle username", "kaggle", "kaggle profile"],
    "scholar id": ["scholar id", "google scholar", "scholar profile"],
    "live project links": [
        "live project links",
        "live projects",
        "project links",
        "portfolio links",
    ],
    # Probable projects
    "project idea": ["project idea", "idea"],
    "author": ["author", "submitted by"],
    "topic": ["topic", "domain"],
}

# Define required and optional columns for each sheet.
SHEET_CONFIG = {
    "Students Info": {
        "aliases": ["students info", "students", "students_info"],
        "required": ["name", "registration number"],
        "optional": ["email", "phone", "file"],
    },
    "Mentors info": {
        "aliases": ["mentors info", "mentors", "mentors_info"],
        "required": ["mentors"],
        "optional": ["email id"],
    },
    "Mentors-projects": {
        "aliases": ["mentors-projects", "mentor projects", "mentor_projects"],
        "required": ["mentors", "project title"],
        "optional": [
            "short profile of the mentor",
            "project abstract",
            "pre-requisites",
            "student's preference - 1",
            "student's preference - 2",
            "student's preference - 3",
            "selected students",
        ],
    },
    "Probable projects": {
        "aliases": ["probable projects", "probable_projects"],
        "required": ["project idea"],
        "optional": ["author", "topic"],
    },
}


def _find_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Worksheet | None:
    """Find a sheet in the workbook using its aliases."""
    config = SHEET_CONFIG[sheet_name]
    for ws_name in wb.sheetnames:
        if _clean_header(ws_name) in config["aliases"]:
            return wb[ws_name]
    return None


def _map_headers(
    ws: Worksheet, sheet_name: str, issues: list[ValidationIssueOut]
) -> dict[str, int] | None:
    """Map canonical column names to their 0-indexed position in the worksheet."""
    is_empty = False
    if ws.max_row < 1 or ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
        is_empty = True

    if is_empty:
        issues.append(
            ValidationIssueOut(
                sheet_name=sheet_name,
                code="sheet.empty",
                severity="error",
                message=f"Sheet '{sheet_name}' is empty.",
            )
        )
        return None

    headers: list[str] = [str(cell.value) if cell.value else "" for cell in ws[1]]
    cleaned_headers = [_clean_header(h) for h in headers]

    col_map = {}
    config = SHEET_CONFIG[sheet_name]
    all_expected = config["required"] + config["optional"]

    for expected in all_expected:
        aliases = HEADER_ALIASES.get(expected, [expected])
        found = False
        for idx, header in enumerate(cleaned_headers):
            if header in aliases:
                col_map[expected] = idx
                found = True
                break

        if not found and expected in config["required"]:
            issues.append(
                ValidationIssueOut(
                    sheet_name=sheet_name,
                    row_number=1,
                    code="sheet.required_column_missing",
                    severity="error",
                    message=f"Missing required column: '{expected}'.",
                )
            )

    # Return None if we are missing any required columns.
    for req in config["required"]:
        if req not in col_map:
            return None

    return col_map


def _get_val(
    row: tuple[Any, ...], col_map: dict[str, int], col_name: str
) -> str | None:
    idx = col_map.get(col_name)
    if idx is None or idx >= len(row):
        return None
    val = row[idx].value
    if val is None:
        return None
    cleaned = str(val).strip()
    if cleaned.lower() in {"n/a", "na", "-"}:
        return None
    return cleaned


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(cell.value is None or str(cell.value).strip() == "" for cell in row)


def _validate_email(email: str | None) -> bool:
    if not email:
        return True
    if "@" not in email:
        return False
    return True


def parse_workbook(file_content: bytes) -> ParsedWorkbook:
    """Parse the workbook and extract all typed rows and validation issues."""
    parsed = ParsedWorkbook()
    issues = parsed.issues

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as e:
        issues.append(
            ValidationIssueOut(
                code="workbook.open_failed",
                severity="error",
                message=f"Failed to open workbook: {e!s}",
                blocking=True,
            )
        )
        return parsed

    # 1. Students Info
    ws_students = _find_sheet(wb, "Students Info")
    if not ws_students:
        issues.append(
            ValidationIssueOut(
                code="sheet.required_missing",
                severity="error",
                message="Missing sheet: 'Students Info'",
            )
        )
    else:
        # Extract Google Drive folder hyperlink
        resumes_url = None
        if ws_students.max_row >= 25:
            cell = ws_students.cell(row=25, column=2)
            if cell.hyperlink and cell.hyperlink.target:
                resumes_url = cell.hyperlink.target
            elif (
                cell.value
                and isinstance(cell.value, str)
                and ("drive.google.com" in cell.value or cell.value.startswith("http"))
            ):
                resumes_url = cell.value.strip()

        if not resumes_url:
            for row in ws_students.iter_rows(min_row=1):
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.target:
                        target = cell.hyperlink.target
                        if "drive.google.com" in target:
                            resumes_url = target
                            break
                    elif (
                        cell.value
                        and isinstance(cell.value, str)
                        and "drive.google.com" in cell.value
                    ):
                        resumes_url = cell.value.strip()
                        break
                if resumes_url:
                    break
        parsed.resumes_url = resumes_url

        col_map = _map_headers(ws_students, "Students Info", issues)
        if col_map:
            seen_regs = set()
            for r_idx, row in enumerate(ws_students.iter_rows(min_row=2), start=2):
                if _is_empty_row(row):
                    continue

                name = _get_val(row, col_map, "name")
                if name and _clean_header(name) in ("student's resume", "all resumes"):
                    continue

                reg = _get_val(row, col_map, "registration number")
                email = _get_val(row, col_map, "email")
                phone = _get_val(row, col_map, "phone")
                file = _get_val(row, col_map, "file")
                github_username = _get_val(row, col_map, "github username") or _get_val(
                    row, col_map, "github"
                )
                leetcode_username = _get_val(
                    row, col_map, "leetcode username"
                ) or _get_val(row, col_map, "leetcode")
                codeforces_username = _get_val(
                    row, col_map, "codeforces username"
                ) or _get_val(row, col_map, "codeforces")
                kaggle_username = _get_val(row, col_map, "kaggle username") or _get_val(
                    row, col_map, "kaggle"
                )
                scholar_id = _get_val(row, col_map, "scholar id") or _get_val(
                    row, col_map, "google scholar"
                )
                live_project_links = _get_val(
                    row, col_map, "live project links"
                ) or _get_val(row, col_map, "live projects")

                raw_data = {
                    "name": name,
                    "registration_number": reg,
                    "email": email,
                    "phone": phone,
                    "file": file,
                    "github_username": github_username,
                    "leetcode_username": leetcode_username,
                    "codeforces_username": codeforces_username,
                    "kaggle_username": kaggle_username,
                    "scholar_id": scholar_id,
                    "live_project_links": live_project_links,
                }

                if not name:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Students Info",
                            row_number=r_idx,
                            column_name="Name",
                            code="candidate.name_missing",
                            severity="error",
                            message="Student name is missing.",
                        )
                    )
                if not reg:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Students Info",
                            row_number=r_idx,
                            column_name="Registration Number",
                            code="candidate.external_id_missing",
                            severity="error",
                            message="Registration number is missing.",
                        )
                    )
                elif reg in seen_regs:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Students Info",
                            row_number=r_idx,
                            column_name="Registration Number",
                            code="candidate.external_id_duplicate",
                            severity="error",
                            message=f"Duplicate registration number: {reg}",
                        )
                    )
                else:
                    seen_regs.add(reg)

                if email and not _validate_email(email):
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Students Info",
                            row_number=r_idx,
                            column_name="Email",
                            code="candidate.email_invalid",
                            severity="warning",
                            message=f"Invalid email format: {email}",
                        )
                    )

                parsed.students.append((r_idx, StudentRow(**raw_data), raw_data))

    # 2. Mentors info
    ws_mentors = _find_sheet(wb, "Mentors info")
    if not ws_mentors:
        issues.append(
            ValidationIssueOut(
                code="sheet.required_missing",
                severity="error",
                message="Missing sheet: 'Mentors info'",
            )
        )
    else:
        col_map = _map_headers(ws_mentors, "Mentors info", issues)
        if col_map:
            for r_idx, row in enumerate(ws_mentors.iter_rows(min_row=2), start=2):
                if _is_empty_row(row):
                    continue

                name = _get_val(row, col_map, "mentors")
                email = _get_val(row, col_map, "email id")

                raw_data = {"name": name, "email": email}

                if not name:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Mentors info",
                            row_number=r_idx,
                            column_name="Mentors",
                            code="mentor.name_missing",
                            severity="error",
                            message="Mentor name is missing.",
                        )
                    )
                if email and not _validate_email(email):
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Mentors info",
                            row_number=r_idx,
                            column_name="email id",
                            code="mentor.email_invalid",
                            severity="warning",
                            message=f"Invalid email format: {email}",
                        )
                    )

                parsed.mentors.append((r_idx, MentorRow(**raw_data), raw_data))

    # 3. Mentors-projects
    ws_projects = _find_sheet(wb, "Mentors-projects")
    if not ws_projects:
        issues.append(
            ValidationIssueOut(
                code="sheet.required_missing",
                severity="error",
                message="Missing sheet: 'Mentors-projects'",
            )
        )
    else:
        col_map = _map_headers(ws_projects, "Mentors-projects", issues)
        if col_map:
            for r_idx, row in enumerate(ws_projects.iter_rows(min_row=2), start=2):
                if _is_empty_row(row):
                    continue

                mentor = _get_val(row, col_map, "mentors")
                profile = _get_val(row, col_map, "short profile of the mentor")
                title = _get_val(row, col_map, "project title")
                abstract = _get_val(row, col_map, "project abstract")
                prereqs = _get_val(row, col_map, "pre-requisites")
                pref1 = _get_val(row, col_map, "student's preference - 1")
                pref2 = _get_val(row, col_map, "student's preference - 2")
                pref3 = _get_val(row, col_map, "student's preference - 3")
                selected = _get_val(row, col_map, "selected students")

                raw_data = {
                    "mentor_name": mentor,
                    "mentor_profile": profile,
                    "title": title,
                    "abstract": abstract,
                    "prerequisites": prereqs,
                    "preference_1": pref1,
                    "preference_2": pref2,
                    "preference_3": pref3,
                    "selected_students": selected,
                }

                if not mentor:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Mentors-projects",
                            row_number=r_idx,
                            column_name="Mentors",
                            code="project.mentor_missing",
                            severity="error",
                            message="Mentor name is missing.",
                        )
                    )
                if not title:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Mentors-projects",
                            row_number=r_idx,
                            column_name="Project Title",
                            code="project.title_missing",
                            severity="error",
                            message="Project title is missing.",
                        )
                    )
                if not abstract:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Mentors-projects",
                            row_number=r_idx,
                            column_name="Project Abstract",
                            code="project.abstract_missing",
                            severity="warning",
                            message="Project abstract is missing.",
                        )
                    )
                if not prereqs:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Mentors-projects",
                            row_number=r_idx,
                            column_name="Pre-requisites",
                            code="project.prerequisites_missing",
                            severity="warning",
                            message="Project prerequisites are missing.",
                        )
                    )

                parsed.mentor_projects.append(
                    (r_idx, MentorProjectRow(**raw_data), raw_data)
                )

    # 4. Probable projects
    ws_probable = _find_sheet(wb, "Probable projects")
    if not ws_probable:
        # Optional sheet, but we log a warning if we want to be strict.
        # The spec doesn't require it to fail.
        pass
    else:
        col_map = _map_headers(ws_probable, "Probable projects", issues)
        if col_map:
            for r_idx, row in enumerate(ws_probable.iter_rows(min_row=2), start=2):
                if _is_empty_row(row):
                    continue

                idea = _get_val(row, col_map, "project idea")
                author = _get_val(row, col_map, "author")
                topic = _get_val(row, col_map, "topic")

                raw_data = {"idea": idea, "author": author, "topic": topic}

                if not idea:
                    issues.append(
                        ValidationIssueOut(
                            sheet_name="Probable projects",
                            row_number=r_idx,
                            column_name="Project Idea",
                            code="project_idea.missing",
                            severity="warning",
                            message="Project idea is missing.",
                        )
                    )

                parsed.probable_projects.append(
                    (r_idx, ProbableProjectRow(**raw_data), raw_data)
                )

    return parsed
